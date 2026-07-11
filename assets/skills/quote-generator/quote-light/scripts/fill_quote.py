"""
견적서 채우기 — assets/quote_template.xlsx 를 복사하고 0.입력 시트에 데이터를 채운다.

사용법:
    python fill_quote.py <input.json> [output.xlsx]

input.json 형식:
{
    "customer": {                       // 공급받는자 (0.입력 시트로 들어감)
        "company": "(주)예시",
        "address": "서울시 강남구 ...",
        "contact_person": "홍길동",
        "phone": "010-1234-5678",
        "email": "hong@example.com",
        "quote_number": "Q-2605-001",
        "issue_date": "2026-05-12",            // 옵션 (없으면 TODAY())
        "validity": "발행일로부터 30일",         // 옵션
        "department": "구매팀"
    },
    "supplier": {                       // 공급자 (1.견적서 시트 우측 박스, 옵션)
        "company": "(주)자사",          // 없으면 템플릿 기본값 "(주)샘플" 유지
        "registration_number": "123-45-67890",
        "ceo": "홍길동",
        "address": "서울시 강남구",
        "department": "영업부",
        "phone": "02-1234-5678"
    },
    "items": [                          // 제품 항목 (최대 5개)
        {
            "product": "제품명",
            "quantity": 100,
            "unit_price": 230000,
            "note": "비고"
        }
    ]
}

미입력 필드는 그냥 비워둠.
"""
import json
import shutil
import subprocess
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook


# 항목 입력 행 매핑 (0.입력 시트)
ITEM_ROW_START = 11
MAX_ITEMS = 5

# 0.입력 시트 — 고객사(공급받는자) 정보 매핑 (병합셀의 좌측 상단 셀 기준)
INPUT_CELLS = {
    "company":        "B5",
    "address":        "E5",    # D5 라벨, E5:F5 병합 입력
    "quote_number":   "H5",    # G5 라벨, H5:K5 병합 입력
    "contact_person": "B6",
    "phone":          "E6",
    "email":          "H6",
    "issue_date":     "B7",
    "validity":       "E7",
    "department":     "H7",
}

# 항목 행의 컬럼:
#   B=제품명(B:E 병합 — B에만 쓰면 됨), F=수량, G=단가, I=비고(I:K 병합 — I에만 쓰면 됨)
ITEM_COLS = {
    "product":    "B",
    "quantity":   "F",
    "unit_price": "G",
    "note":       "I",
}

# 1.견적서 시트 — 공급자 정보 매핑 (H5~H10, 병합 H:J)
# 견적서 시트는 평소엔 템플릿 하드코딩 값을 보여주지만, supplier 정보가 들어오면 덮어씀
SUPPLIER_CELLS = {
    "company":             "H5",
    "registration_number": "H6",
    "ceo":                 "H7",
    "address":             "H8",
    "department":          "H9",
    "phone":               "H10",
}


def _coerce_number(value, *, field: str, item_index: int):
    """수량/단가를 숫자로 변환. 변환 실패 시 경고하고 빈칸 처리(잘못된 값을 Excel에 넘기지 않음)."""
    if isinstance(value, bool):  # True/False가 1/0으로 새는 것 방지
        print(f"⚠️  항목 {item_index}의 {field} 값이 숫자가 아닙니다(boolean): {value!r}. "
              f"비워둡니다. 0.입력 시트에서 직접 채우세요.", file=sys.stderr)
        return None
    if isinstance(value, (int, float)):
        return value
    # 문자열 등: "35,000", "35000원" 같은 흔한 표기 정리 후 시도
    cleaned = str(value).replace(",", "").replace("원", "").strip()
    try:
        num = float(cleaned)
        return int(num) if num.is_integer() else num
    except (ValueError, TypeError):
        print(f"⚠️  항목 {item_index}의 {field} 값이 숫자가 아닙니다: {value!r}. "
              f"비워둡니다. 0.입력 시트 해당 셀에 직접 숫자를 입력하면 합계가 자동 계산됩니다.",
              file=sys.stderr)
        return None


def fill_quote(template_path: str, data: dict, output_path: str):
    # 템플릿 복사
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)

    # ===== 1. 고객사 정보 (0.입력 시트) =====
    ws_input = wb["0.입력"]
    customer = data.get("customer", {})
    for key, cell in INPUT_CELLS.items():
        value = customer.get(key)
        if value is None or value == "":
            continue
        # 발행일은 YYYY-MM-DD면 datetime으로, 아니면 경고하고 문자열 그대로 둠
        if key == "issue_date":
            try:
                value = datetime.strptime(str(value), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                print(f"⚠️  발행일 형식이 YYYY-MM-DD가 아닙니다: {value!r}. "
                      f"날짜로 인식되지 않고 입력값 그대로 들어갑니다. "
                      f"날짜 정렬/표기가 필요하면 0.입력 시트 B7을 수정하세요.",
                      file=sys.stderr)
        ws_input[cell] = value

    # ===== 2. 제품 항목 (0.입력 시트) =====
    # 부분 항목(수량/단가 누락)도 허용한다. 빈 셀은 비운 채 저장되며,
    # 0.입력 시트에서 나중에 채우면 1.견적서 시트와 합계가 자동 반영된다.
    items = data.get("items", [])
    if len(items) > MAX_ITEMS:
        print(f"⚠️  항목이 {MAX_ITEMS}개를 초과합니다. 앞 {MAX_ITEMS}개만 입력합니다.",
              file=sys.stderr)
        items = items[:MAX_ITEMS]

    for i, item in enumerate(items):
        row = ITEM_ROW_START + i
        for key, col in ITEM_COLS.items():
            value = item.get(key)
            if value is None or value == "":
                continue  # 부분 항목 허용 — 빈 채로 둔다
            if key in ("quantity", "unit_price"):
                value = _coerce_number(value, field=key, item_index=i + 1)
                if value is None:
                    continue  # 숫자로 못 바꾸면 빈칸 — 잘못된 값을 Excel에 넣지 않음
            ws_input[f"{col}{row}"] = value

    # ===== 3. 공급자 정보 (1.견적서 시트, 옵션) =====
    supplier = data.get("supplier", {})
    if supplier:
        ws_quote = wb["1.견적서"]
        for key, cell in SUPPLIER_CELLS.items():
            value = supplier.get(key)
            if value is None or value == "":
                continue
            ws_quote[cell] = value

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python fill_quote.py <input.json> [output.xlsx]",
              file=sys.stderr)
        sys.exit(1)

    input_json = Path(sys.argv[1])
    output_path = sys.argv[2] if len(sys.argv) > 2 else "quote_output.xlsx"

    with open(input_json, encoding="utf-8") as f:
        data = json.load(f)

    # 템플릿은 스크립트와 같은 폴더의 ../assets/quote_template.xlsx
    script_dir = Path(__file__).resolve().parent
    template = script_dir.parent / "assets" / "quote_template.xlsx"

    # 템플릿이 없거나 손상돼 열리지 않으면 create_template.py로 자동 재생성 후 재시도
    needs_rebuild = not template.exists()
    if not needs_rebuild:
        try:
            load_workbook(str(template)).close()
        except Exception as e:
            print(f"⚠️  템플릿을 열 수 없습니다({e}). 재생성합니다: {template}",
                  file=sys.stderr)
            needs_rebuild = True

    if needs_rebuild:
        creator = script_dir / "create_template.py"
        if not creator.exists():
            print(f"❌ 템플릿도 없고 생성기도 없습니다: {creator}", file=sys.stderr)
            sys.exit(1)
        print(f"ℹ️  템플릿이 없어 재생성합니다: {creator} → {template}",
              file=sys.stderr)
        template.parent.mkdir(parents=True, exist_ok=True)
        result = subprocess.run(
            [sys.executable, str(creator), str(template)],
            capture_output=True, text=True,
        )
        if result.returncode != 0 or not template.exists():
            print(f"❌ 템플릿 재생성 실패:\n{result.stderr}", file=sys.stderr)
            sys.exit(1)

    fill_quote(str(template), data, output_path)
    print(f"✅ 견적서 생성 완료: {output_path}")


if __name__ == "__main__":
    main()
