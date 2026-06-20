"""
공유용 가벼운 견적서 템플릿 생성기.

기존 자동_견적서_템플릿_fn_v1_1.xlsx 의 디자인을 참고하되,
- #REF! 깨진 수식 제거
- 내부용 단가 자동계산 로직 제거 (단가는 직접 입력)
- 입력 시트 + 견적서 시트 연동 수식은 유지 (B5, 모델/수량/단가 등 → 견적서)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ===== 색상 =====
COLOR_HEADER_DARK = "344E7A"   # 표 헤더 (진한 남색)
COLOR_HEADER_LIGHT = "D9E1F2"  # 섹션 헤더 (연한 파랑)
COLOR_LABEL_GRAY = "F2F2F2"    # 라벨 셀
COLOR_INPUT_BLUE = "DDEBF7"    # 직접입력 (파란셀)
COLOR_DROPDOWN_PINK = "FCE4D6" # 드롭다운 (분홍셀)
COLOR_AUTO_GRAY = "EDEDED"     # 자동계산 (회색셀)
COLOR_WHITE = "FFFFFF"

# ===== 폰트 =====
FONT_NAME = "맑은 고딕"

def make_border(style="thin", color="808080"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

THIN_BORDER = make_border("thin", "808080")
MEDIUM_BORDER = make_border("medium", "000000")

def style_cell(cell, *, bold=False, size=10, color="000000",
               fill_color=None, h_align="center", v_align="center",
               border=THIN_BORDER, wrap=True):
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
    if fill_color:
        cell.fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    cell.border = border


def build_input_sheet(ws):
    """0.입력 시트 — 고객사 정보 + 모델 입력 (최대 5개)."""

    # 열 너비
    widths = {"A": 12, "B": 18, "C": 12, "D": 14, "E": 16, "F": 12,
              "G": 14, "H": 18, "I": 14, "J": 10, "K": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ===== Row 1: 타이틀 =====
    ws.merge_cells("A1:K1")
    ws["A1"] = "견적 입력 시트"
    style_cell(ws["A1"], bold=True, size=16, fill_color=COLOR_HEADER_DARK,
               color=COLOR_WHITE, border=None)
    ws.row_dimensions[1].height = 30

    # ===== Row 2: 범례 =====
    ws.merge_cells("A2:K2")
    ws["A2"] = "🔵 파란셀=직접입력   ⬜ 회색셀=자동계산"
    style_cell(ws["A2"], size=9, fill_color=COLOR_WHITE, h_align="center",
               border=None)
    ws.row_dimensions[2].height = 20

    # ===== Row 4: 고객사 정보 헤더 =====
    ws.merge_cells("A4:K4")
    ws["A4"] = "■ 고객사 정보"
    style_cell(ws["A4"], bold=True, size=11, fill_color=COLOR_HEADER_LIGHT,
               h_align="left", border=None)
    ws.row_dimensions[4].height = 22

    # ===== Row 5-7: 고객사 정보 입력 (3 x 3 그리드 형태) =====
    # 라벨/입력 쌍을 3개씩 한 행에 배치
    customer_layout = [
        # (row, [(label_col, label, input_col_start, input_col_end)])
        (5, [("A", "법 인 명", "B", "C"),
             ("D", "주     소", "E", "F"),
             ("G", "견 적 번 호", "H", "K")]),
        (6, [("A", "담 당 자", "B", "C"),
             ("D", "연 락 처", "E", "F"),
             ("G", "이  메  일", "H", "K")]),
        (7, [("A", "발 행 일", "B", "C"),
             ("D", "유 효 기 간", "E", "F"),
             ("G", "담 당 부 서", "H", "K")]),
    ]

    for row, items in customer_layout:
        ws.row_dimensions[row].height = 24
        for label_col, label, in_start, in_end in items:
            # 라벨 셀
            cell = ws[f"{label_col}{row}"]
            cell.value = label
            style_cell(cell, bold=True, size=10, fill_color=COLOR_LABEL_GRAY)
            # 입력 셀 병합 + 스타일
            merge_range = f"{in_start}{row}:{in_end}{row}"
            ws.merge_cells(merge_range)
            input_cell = ws[f"{in_start}{row}"]
            style_cell(input_cell, fill_color=COLOR_INPUT_BLUE, h_align="left")

    # 기본값
    ws["B7"] = "=TODAY()"
    ws["B7"].number_format = "yyyy-mm-dd"
    ws["E7"] = "발행일로부터 30일"
    style_cell(ws["E7"], fill_color=COLOR_AUTO_GRAY, h_align="left")

    # ===== Row 9: 견적 모델 설정 헤더 =====
    ws.merge_cells("A9:K9")
    ws["A9"] = "■ 견적 항목 (최대 5개 — 파란셀=직접입력)"
    style_cell(ws["A9"], bold=True, size=11, fill_color=COLOR_HEADER_LIGHT,
               h_align="left", border=None)
    ws.row_dimensions[9].height = 22

    # ===== Row 10: 항목 표 헤더 =====
    # 컬럼 단순화: #(A) | 제품명(B:E) | 수량(F) | 단가(G) | 공급가액(H) | 비고(I:K)
    ws.merge_cells("B10:E10")
    ws.merge_cells("I10:K10")
    headers_map = {
        "A10": "#",
        "B10": "제품명",
        "F10": "수량",
        "G10": "단가\n(원)",
        "H10": "공급가액\n(자동)",
        "I10": "비고",
    }
    for cell_ref, h in headers_map.items():
        ws[cell_ref] = h
        style_cell(ws[cell_ref], bold=True, size=10,
                   fill_color=COLOR_HEADER_DARK, color=COLOR_WHITE)
    # 병합 영역의 나머지 셀에도 헤더 스타일 적용
    for ref in ["B10:E10", "I10:K10"]:
        for r in ws[ref]:
            for c in r:
                c.border = THIN_BORDER
                c.fill = PatternFill("solid", start_color=COLOR_HEADER_DARK,
                                     end_color=COLOR_HEADER_DARK)
    ws.row_dimensions[10].height = 32

    # ===== Row 11-15: 항목 입력 5행 =====
    for i, row in enumerate(range(11, 16), start=1):
        ws.row_dimensions[row].height = 22

        # A: # (순번, 자동, 회색)
        ws.cell(row=row, column=1, value=i)
        style_cell(ws.cell(row=row, column=1), fill_color=COLOR_AUTO_GRAY)

        # B:E 제품명 (직접입력, 파랑) — 병합
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        style_cell(ws.cell(row=row, column=2),
                   fill_color=COLOR_INPUT_BLUE, h_align="left")
        for col_idx in range(3, 6):
            style_cell(ws.cell(row=row, column=col_idx),
                       fill_color=COLOR_INPUT_BLUE)

        # F: 수량 (직접입력, 파랑)
        style_cell(ws.cell(row=row, column=6), fill_color=COLOR_INPUT_BLUE)
        ws.cell(row=row, column=6).number_format = "#,##0"

        # G: 단가 (직접입력, 파랑)
        style_cell(ws.cell(row=row, column=7), fill_color=COLOR_INPUT_BLUE)
        ws.cell(row=row, column=7).number_format = "#,##0"

        # H: 공급가액 = 수량 × 단가 (자동, 회색)
        formula = (
            f'=IF(OR(F{row}="",G{row}="",F{row}=0,G{row}=0),"",'
            f'F{row}*G{row})'
        )
        ws.cell(row=row, column=8, value=formula)
        style_cell(ws.cell(row=row, column=8), fill_color=COLOR_AUTO_GRAY)
        ws.cell(row=row, column=8).number_format = "#,##0"

        # I:K 비고 (병합, 파랑)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
        style_cell(ws.cell(row=row, column=9), fill_color=COLOR_INPUT_BLUE,
                   h_align="left")
        for col_idx in range(10, 12):
            style_cell(ws.cell(row=row, column=col_idx),
                       fill_color=COLOR_INPUT_BLUE)

    # ===== Row 17: 합계 =====
    ws.row_dimensions[17].height = 26
    ws.merge_cells("A17:G17")
    ws["A17"] = "합 계 (VAT 별도)"
    style_cell(ws["A17"], bold=True, size=11, fill_color=COLOR_HEADER_LIGHT,
               h_align="right")
    ws["H17"] = "=IFERROR(SUM(H11:H15),0)"
    ws["H17"].number_format = "#,##0"
    style_cell(ws["H17"], bold=True, size=11, fill_color=COLOR_HEADER_LIGHT)
    ws.merge_cells("I17:K17")
    style_cell(ws["I17"], fill_color=COLOR_HEADER_LIGHT)
    for col_idx in range(10, 12):
        style_cell(ws.cell(row=17, column=col_idx),
                   fill_color=COLOR_HEADER_LIGHT)

    # ===== Row 19: 안내 =====
    ws.merge_cells("A19:K19")
    ws["A19"] = ("※ 공급가액 = 수량 × 단가 (자동계산).  "
                "제품명, 수량, 단가는 직접 입력하세요.")
    style_cell(ws["A19"], size=9, fill_color=COLOR_WHITE, h_align="left",
               border=None)
    ws.row_dimensions[19].height = 24


def build_quote_sheet(ws):
    """1.견적서 시트 — 0.입력 시트 데이터를 참조해 자동 생성되는 견적서."""

    # 열 너비 (기존 템플릿 참고)
    widths = {"A": 2, "B": 10, "C": 11, "D": 13, "E": 13, "F": 8,
              "G": 12, "H": 10, "I": 12, "J": 14, "K": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ===== Row 3: 타이틀 =====
    ws.row_dimensions[3].height = 40
    ws.merge_cells("B3:J3")
    ws["B3"] = "견  적  서"
    ws["B3"].font = Font(name=FONT_NAME, size=24, bold=True)
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # ===== Row 5-10: 공급받는자 / 공급자 박스 =====
    # 공급받는자 (왼쪽): B5:E10
    ws.merge_cells("B5:B10")
    ws["B5"] = "공\n급\n받\n는\n자"
    style_cell(ws["B5"], bold=True, size=11, fill_color=COLOR_LABEL_GRAY,
               border=MEDIUM_BORDER)

    # 공급자 (오른쪽): F5:J10
    ws.merge_cells("F5:F10")
    ws["F5"] = "공\n급\n자"
    style_cell(ws["F5"], bold=True, size=11, fill_color=COLOR_LABEL_GRAY,
               border=MEDIUM_BORDER)

    # 라벨/값 행 구성
    # (row, 좌측 라벨, 좌측 값 수식, 우측 라벨, 우측 값)
    # 공급받는자: 0.입력 시트 참조 (빈 셀일 때 0 표시 방지 위해 IF 처리)
    # 공급자: 법인명만 (주)샘플로 두고 나머지는 빈칸 (사용자가 직접 채움)
    info_rows = [
        (5, "법 인 명",    '=IF(\'0.입력\'!B5="","",\'0.입력\'!B5)',
            "법 인 명",  "(주)샘플"),
        (6, "주     소",   '=IF(\'0.입력\'!E5="","",\'0.입력\'!E5)',
            "등록번호",  ""),
        (7, "담 당 자",    '=IF(\'0.입력\'!B6="","",\'0.입력\'!B6)',
            "대 표 자",  ""),
        (8, "연 락 처",    '=IF(\'0.입력\'!E6="","",\'0.입력\'!E6)',
            "주     소",  ""),
        (9, "발 행 일",    '=IF(\'0.입력\'!B7="","",\'0.입력\'!B7)',
            "담당부서",  ""),
        (10, "유 효 기 간", '=IF(\'0.입력\'!E7="","",\'0.입력\'!E7)',
            "연 락 처",  ""),
    ]

    for row, l_label, l_value, r_label, r_value in info_rows:
        ws.row_dimensions[row].height = 22
        # 좌측 라벨
        ws[f"C{row}"] = l_label
        style_cell(ws[f"C{row}"], bold=True, size=10,
                   fill_color=COLOR_LABEL_GRAY)
        # 좌측 값 (D:E 병합)
        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"] = l_value
        style_cell(ws[f"D{row}"], size=10, h_align="left")
        if row == 9:  # 발행일
            ws[f"D{row}"].number_format = "yyyy-mm-dd"

        # 우측 라벨
        ws[f"G{row}"] = r_label
        style_cell(ws[f"G{row}"], bold=True, size=10,
                   fill_color=COLOR_LABEL_GRAY)
        # 우측 값 (H:J 병합)
        ws.merge_cells(f"H{row}:J{row}")
        ws[f"H{row}"] = r_value
        style_cell(ws[f"H{row}"], size=10, h_align="left")

    # ===== Row 13-14: 수신 + 인사 =====
    ws.row_dimensions[13].height = 24
    ws.merge_cells("B13:J13")
    # 회사명이 비어있으면 빈 문자열, 있으면 "수신: 회사명 귀중"
    ws["B13"] = ('=IF(\'0.입력\'!B5="","",'
                 '" 수  신 : "&\'0.입력\'!B5&"  귀중 ")')
    ws["B13"].font = Font(name=FONT_NAME, size=12, bold=True)
    ws["B13"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[14].height = 20
    ws.merge_cells("B14:J14")
    ws["B14"] = "  상기와 같이 견적서를 제출합니다."
    ws["B14"].font = Font(name=FONT_NAME, size=10)
    ws["B14"].alignment = Alignment(horizontal="left", vertical="center")

    # ===== Row 16: 견적 상세 내역 헤더 =====
    ws.row_dimensions[16].height = 24
    ws.merge_cells("B16:J16")
    ws["B16"] = "■ 견적 상세 내역"
    style_cell(ws["B16"], bold=True, size=11, fill_color=COLOR_HEADER_LIGHT,
               h_align="left", border=None)

    # ===== Row 18: 표 헤더 =====
    # 컬럼: 항목(B:D) | 수량(E) | 단가(F) | 공급가액(G:H) | 비고(I:J)
    ws.row_dimensions[18].height = 28
    header_cells = [
        ("B18:D18", "항    목"),
        ("E18", "수량"),
        ("F18", "단가"),
        ("G18:H18", "공급가액"),
        ("I18:J18", "비고"),
    ]
    for ref, val in header_cells:
        if ":" in ref:
            ws.merge_cells(ref)
            first = ref.split(":")[0]
        else:
            first = ref
        ws[first] = val
        style_cell(ws[first], bold=True, size=10, fill_color=COLOR_HEADER_DARK,
                   color=COLOR_WHITE)
        # 병합 영역 전체에 border + 헤더색
        if ":" in ref:
            for r in ws[ref]:
                for c in r:
                    c.border = THIN_BORDER
                    c.fill = PatternFill("solid", start_color=COLOR_HEADER_DARK,
                                         end_color=COLOR_HEADER_DARK)

    # ===== Row 19-23: 항목 5행 (입력 시트 11-15행에서 참조) =====
    # 입력 시트 컬럼: B=제품명, F=수량, G=단가, H=공급가액, I=비고
    for i, row in enumerate(range(19, 24)):
        input_row = 11 + i  # 0.입력 시트의 대응 행
        ws.row_dimensions[row].height = 24

        # B:D 항목 (제품명)
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = f'=IF(\'0.입력\'!B{input_row}="","",\'0.입력\'!B{input_row})'
        style_cell(ws[f"B{row}"], size=10, h_align="left")
        for col in ["C", "D"]:
            style_cell(ws[f"{col}{row}"], size=10)

        # E: 수량
        ws[f"E{row}"] = f'=IF(\'0.입력\'!F{input_row}="","",\'0.입력\'!F{input_row})'
        ws[f"E{row}"].number_format = "#,##0"
        style_cell(ws[f"E{row}"], size=10)

        # F: 단가
        ws[f"F{row}"] = f'=IF(\'0.입력\'!G{input_row}="","",\'0.입력\'!G{input_row})'
        ws[f"F{row}"].number_format = "#,##0"
        style_cell(ws[f"F{row}"], size=10)

        # G:H 공급가액
        ws.merge_cells(f"G{row}:H{row}")
        ws[f"G{row}"] = f'=IF(\'0.입력\'!H{input_row}="","",\'0.입력\'!H{input_row})'
        ws[f"G{row}"].number_format = "#,##0"
        style_cell(ws[f"G{row}"], size=10)
        style_cell(ws[f"H{row}"], size=10)

        # I:J 비고
        ws.merge_cells(f"I{row}:J{row}")
        ws[f"I{row}"] = f'=IF(\'0.입력\'!I{input_row}="","",\'0.입력\'!I{input_row})'
        style_cell(ws[f"I{row}"], size=10, h_align="left")
        style_cell(ws[f"J{row}"], size=10)

    # ===== Row 25: 합계 =====
    ws.row_dimensions[25].height = 28
    ws.merge_cells("B25:F25")
    ws["B25"] = "합     계  (VAT 별도)"
    style_cell(ws["B25"], bold=True, size=12, fill_color=COLOR_HEADER_LIGHT,
               h_align="right")

    ws.merge_cells("G25:H25")
    ws["G25"] = "=IFERROR(SUM(G19:G23),0)"
    ws["G25"].number_format = "#,##0"
    style_cell(ws["G25"], bold=True, size=12, fill_color=COLOR_HEADER_LIGHT)

    ws.merge_cells("I25:J25")
    ws["I25"] = "원"
    style_cell(ws["I25"], bold=True, size=12, fill_color=COLOR_HEADER_LIGHT)

    # ===== Row 27-30: 안내 문구 =====
    notes = [
        "※ 본 견적서의 유효기간은 발행일로부터 30일이며, 부가가치세는 별도입니다.",
        "※ 단가 및 공급조건은 협의에 따라 변경될 수 있습니다.",
    ]
    for i, note in enumerate(notes):
        row = 27 + i
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"B{row}:J{row}")
        ws[f"B{row}"] = note
        ws[f"B{row}"].font = Font(name=FONT_NAME, size=9, color="555555")
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")


def main(output_path):
    wb = Workbook()

    # 첫 시트 이름 변경
    ws_input = wb.active
    ws_input.title = "0.입력"
    build_input_sheet(ws_input)

    # 견적서 시트 추가
    ws_quote = wb.create_sheet("1.견적서")
    build_quote_sheet(ws_quote)

    # 인쇄 설정 (견적서 시트)
    ws_quote.page_setup.orientation = ws_quote.ORIENTATION_PORTRAIT
    ws_quote.page_setup.paperSize = ws_quote.PAPERSIZE_A4
    ws_quote.page_setup.fitToWidth = 1
    ws_quote.page_setup.fitToHeight = 1
    ws_quote.sheet_properties.pageSetUpPr.fitToPage = True
    ws_quote.page_margins.left = 0.5
    ws_quote.page_margins.right = 0.5
    ws_quote.page_margins.top = 0.5
    ws_quote.page_margins.bottom = 0.5

    # 기본 시트는 입력 시트
    wb.active = 0

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "quote_template.xlsx"
    main(output)
